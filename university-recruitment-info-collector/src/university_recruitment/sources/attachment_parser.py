"""Attachment parsing helpers for recruitment notices.

This module is used as a fallback when a notice body is too short (e.g. "见附件1").
"""

from __future__ import annotations

import csv
import io
import logging
import re
from typing import Iterable

import httpx
from bs4 import BeautifulSoup

from university_recruitment.config import LLM_MAX_ATTACHMENT_CHARS

logger = logging.getLogger(__name__)

_ATTACHMENT_EXTENSIONS = (".xlsx", ".xls", ".csv", ".pdf", ".docx", ".doc")
_ATTACHMENT_HINT_RE = re.compile(r"(见附件|详见附件|附件\d*|附表\d*)")


def looks_like_attachment_notice(text: str) -> bool:
    """Return True when notice text likely points to attachment content."""
    normalized = (text or "").strip()
    if len(normalized) > 100:
        return False
    return bool(_ATTACHMENT_HINT_RE.search(normalized))


def extract_attachment_urls_from_html(html: str, page_url: str) -> list[tuple[str, str]]:
    """Extract candidate attachment URLs from HTML.

    Returns list[(filename, absolute_url)].
    """
    soup = BeautifulSoup(html, "html.parser")
    base = httpx.URL(page_url)
    attachments: list[tuple[str, str]] = []
    seen: set[str] = set()

    for link in soup.find_all("a", href=True):
        href = (link.get("href") or "").strip()
        if not href or href.startswith(("javascript:", "mailto:")):
            continue
        text = (link.get_text(" ", strip=True) or "").strip()
        lower = href.lower()
        if not any(ext in lower or ext in text.lower() for ext in _ATTACHMENT_EXTENSIONS):
            continue

        try:
            full = str(base.join(href))
        except Exception:
            continue
        if full in seen:
            continue
        seen.add(full)
        filename = text or full.split("/")[-1]
        attachments.append((filename, full))

    return attachments


def parse_attachment_bytes(filename: str, content: bytes, max_chars: int = LLM_MAX_ATTACHMENT_CHARS) -> str:
    """Parse attachment bytes into plain text.

    Unsupported formats return an empty string.
    """
    lower = filename.lower()
    text = ""

    if lower.endswith(".csv"):
        text = _parse_csv(content)
    elif lower.endswith((".xlsx", ".xls")):
        text = _parse_excel(content)
    elif lower.endswith(".docx"):
        text = _parse_docx(content)
    elif lower.endswith(".pdf"):
        text = _parse_pdf(content)

    if not text:
        return ""
    if len(text) > max_chars:
        return text[:max_chars]
    return text


def build_attachment_augmented_text(
    body_text: str,
    notice_url: str,
    max_attachments: int = 3,
    timeout: float = 20.0,
    max_chars: int = LLM_MAX_ATTACHMENT_CHARS,
) -> tuple[str, list[str]]:
    """Fetch notice attachments and append parsed text to body_text.

    Returns (augmented_text, warnings).
    """
    warnings: list[str] = []
    body = (body_text or "").strip()

    try:
        resp = httpx.get(notice_url, follow_redirects=True, timeout=timeout)
        resp.raise_for_status()
    except Exception as exc:
        return body, [f"attachment detail fetch failed: {exc}"]

    attachments = extract_attachment_urls_from_html(resp.text, notice_url)
    if not attachments:
        return body, ["no attachment links found"]

    blocks: list[str] = []
    for idx, (filename, url) in enumerate(attachments[:max_attachments], start=1):
        try:
            file_resp = httpx.get(url, follow_redirects=True, timeout=timeout)
            file_resp.raise_for_status()
            parse_name = filename
            if not any(parse_name.lower().endswith(ext) for ext in _ATTACHMENT_EXTENSIONS):
                parse_name = url.split("/")[-1] or parse_name
            parsed = parse_attachment_bytes(parse_name, file_resp.content, max_chars=max_chars)
            if parsed:
                blocks.append(f"[附件{idx}: {filename}]\n{parsed}")
            else:
                warnings.append(f"attachment parsed empty: {filename}")
        except Exception as exc:
            warnings.append(f"attachment fetch/parse failed: {filename} ({exc})")

    if not blocks:
        return body, warnings or ["all attachments empty"]

    merged = body
    if merged:
        merged += "\n\n"
    merged += "\n\n".join(blocks)
    if len(merged) > max_chars:
        merged = merged[:max_chars]
    return merged, warnings


def prepare_llm_input_text(
    body_text: str,
    notice_url: str,
    min_chars: int = 60,
    max_attachments: int = 3,
    timeout: float = 20.0,
    max_chars: int = LLM_MAX_ATTACHMENT_CHARS,
) -> tuple[str, bool, list[str]]:
    """Prepare LLM input text, augmenting short attachment-style notices."""
    normalized = (body_text or "").strip()
    if len(normalized) >= min_chars:
        return normalized, False, []
    if not looks_like_attachment_notice(normalized):
        return normalized, False, []

    augmented, warnings = build_attachment_augmented_text(
        body_text=normalized,
        notice_url=notice_url,
        max_attachments=max_attachments,
        timeout=timeout,
        max_chars=max_chars,
    )
    if len((augmented or "").strip()) >= min_chars:
        return augmented.strip(), True, warnings
    return normalized, False, warnings


def _parse_csv(content: bytes) -> str:
    text_data = _decode_text(content)
    if not text_data:
        return ""
    reader = csv.reader(io.StringIO(text_data))
    lines: list[str] = []
    for i, row in enumerate(reader):
        lines.append(" | ".join(str(c).strip() for c in row if str(c).strip()))
        if i >= 800:
            break
    return "\n".join(line for line in lines if line).strip()


def _parse_excel(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return ""

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return ""

    chunks: list[str] = []
    row_count = 0
    for ws in wb.worksheets:
        chunks.append(f"[Sheet] {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if not cells:
                continue
            chunks.append(" | ".join(cells))
            row_count += 1
            if row_count >= 2000:
                break
        if row_count >= 2000:
            break

    return "\n".join(chunks).strip()


def _parse_docx(content: bytes) -> str:
    try:
        from docx import Document
    except Exception:
        return ""

    try:
        doc = Document(io.BytesIO(content))
    except Exception:
        return ""

    lines = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
    return "\n".join(lines).strip()


def _parse_pdf(content: bytes) -> str:
    try:
        from pypdf import PdfReader
    except Exception:
        return ""

    try:
        reader = PdfReader(io.BytesIO(content))
    except Exception:
        return ""

    lines: list[str] = []
    for page in reader.pages[:20]:
        txt = page.extract_text() or ""
        cleaned = txt.strip()
        if cleaned:
            lines.append(cleaned)
    return "\n\n".join(lines).strip()


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            return content.decode(encoding)
        except Exception:
            continue
    return ""
